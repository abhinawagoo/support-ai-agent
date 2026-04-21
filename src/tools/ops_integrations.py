from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def supabase_insert_json(
    *,
    supabase_url: str,
    supabase_key: str,
    table: str,
    payload: dict[str, Any],
) -> dict[str, Any]:
    url = f"{supabase_url.rstrip('/')}/rest/v1/{table}"
    headers = {
        "apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    resp = requests.post(url, headers=headers, json=payload, timeout=20)
    resp.raise_for_status()
    data = resp.json()
    if isinstance(data, list) and data:
        return data[0]
    return {"ok": True}


def telegram_send_message(*, bot_token: str, chat_id: str, text: str) -> None:
    url = f"https://api.telegram.org/bot{bot_token}/sendMessage"
    resp = requests.post(url, json={"chat_id": chat_id, "text": text}, timeout=20)
    resp.raise_for_status()


def slack_send_message(*, webhook_url: str, text: str) -> None:
    resp = requests.post(webhook_url, json={"text": text}, timeout=20)
    resp.raise_for_status()


def update_excel_local(*, file_path: str, row: dict[str, Any]) -> None:
    from openpyxl import Workbook, load_workbook

    fp = Path(file_path)
    if fp.exists():
        wb = load_workbook(fp)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Reports"
    headers = list(row.keys())
    if ws.max_row == 1 and ws.cell(1, 1).value is None:
        for idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=idx, value=h)
    ws.append([row.get(h) for h in headers])
    wb.save(fp)


def append_google_sheet(
    *,
    service_account_json: str,
    spreadsheet_id: str,
    sheet_name: str,
    row: dict[str, Any],
) -> None:
    import gspread

    gc = gspread.service_account(filename=service_account_json)
    sh = gc.open_by_key(spreadsheet_id)
    ws = sh.worksheet(sheet_name)
    ws.append_row([str(v) if v is not None else "" for v in row.values()], value_input_option="USER_ENTERED")


def read_google_doc_text(*, service_account_json: str, doc_id: str) -> str:
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build

    scopes = ["https://www.googleapis.com/auth/documents.readonly"]
    creds = Credentials.from_service_account_file(service_account_json, scopes=scopes)
    service = build("docs", "v1", credentials=creds, cache_discovery=False)
    doc = service.documents().get(documentId=doc_id).execute()
    body = doc.get("body", {}).get("content", [])
    chunks: list[str] = []
    for elem in body:
        para = elem.get("paragraph")
        if not para:
            continue
        for pe in para.get("elements", []):
            tr = pe.get("textRun")
            if tr and "content" in tr:
                chunks.append(tr["content"])
    return "".join(chunks).strip()


def safe_json_dumps(data: Any) -> str:
    return json.dumps(data, ensure_ascii=False, indent=2)
